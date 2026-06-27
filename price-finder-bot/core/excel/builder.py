"""Builds the Persian RTL Excel output.

Layout (right-to-left):
  - Summary header row: min / max / average price
  - Columns: row#, product name, price (sorted low->high), store,
             stock availability, seller rating, hyperlinked purchase link
"""
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import settings
from core.aggregator.aggregate import price_summary
from core.models import ProductResult

_HEADERS = ["ردیف", "نام محصول", "قیمت (تومان)", "فروشگاه",
            "موجودی", "امتیاز فروشنده", "لینک خرید"]

_FONT = "Tahoma"
_HEADER_FILL = PatternFill("solid", start_color="2E5A88")
_SUMMARY_FILL = PatternFill("solid", start_color="DCE6F1")
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_excel(query: str, results: List[ProductResult]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "نتایج جستجو"
    ws.sheet_view.rightToLeft = True

    summary = price_summary(results)

    # --- Row 1: query + summary ---
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = (f"جستجو: «{query}»   |   کمترین: {_fmt(summary['min'])}   "
                   f"بیشترین: {_fmt(summary['max'])}   میانگین: {_fmt(summary['avg'])}   "
                   f"|   تعداد: {summary['count']}")
    title.font = Font(name=_FONT, bold=True, size=12, color="1F3864")
    title.fill = _SUMMARY_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # --- Row 2: column headers ---
    for col, head in enumerate(_HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=head)
        c.font = Font(name=_FONT, bold=True, color="FFFFFF")
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
    ws.row_dimensions[2].height = 22

    # --- Data rows (already sorted low->high by aggregator) ---
    for i, r in enumerate(results, start=1):
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r.name)
        ws.cell(row=row, column=3, value=r.price)
        ws.cell(row=row, column=4, value=r.store)
        ws.cell(row=row, column=5, value="موجود" if r.in_stock else "ناموجود")
        ws.cell(row=row, column=6, value=r.rating if r.rating is not None else "-")

        link_cell = ws.cell(row=row, column=7, value="مشاهده و خرید")
        if r.link:
            link_cell.hyperlink = r.link
            link_cell.font = Font(name=_FONT, color="0563C1", underline="single")

        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.font.name != _FONT and col != 7:
                cell.font = Font(name=_FONT)
        ws.cell(row=row, column=3).number_format = "#,##0"

    widths = [8, 45, 16, 22, 12, 16, 18]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = settings.OUTPUT_DIR / f"prices_{ts}.xlsx"
    wb.save(out)
    return out


def _fmt(v):
    return f"{v:,}" if isinstance(v, int) else "—"
